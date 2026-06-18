# =============================================================================
# Crea il file Excel di input per il launcher di test
# Struttura identica a DGS_forecasting_short.m
# Esegui questo script UNA VOLTA per creare il file di test
# =============================================================================
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import numpy as np

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Input"

# --- Stili ---
header_fill  = PatternFill("solid", fgColor="2F5496")
header_font  = Font(color="FFFFFF", bold=True)
series_fills = [
    PatternFill("solid", fgColor="DEEAF1"),  # azzurro chiaro
    PatternFill("solid", fgColor="E2EFDA"),  # verde chiaro
    PatternFill("solid", fgColor="FFF2CC"),  # giallo chiaro
    PatternFill("solid", fgColor="FCE4D6"),  # arancio chiaro
]
thin = Side(style='thin', color='AAAAAA')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- Header riga 1 ---
ws.merge_cells("A1:B1")
ws["A1"] = "DGS Forecast Engine — Test Launcher"
ws["A1"].font = Font(bold=True, size=13, color="2F5496")
ws["A1"].alignment = Alignment(horizontal="center")

# --- Header riga 3 ---
headers = ["SKU_ID", "Mese", "S1_trend", "S2_stagionale", "S3_intermittente", "S4_rumore"]
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=c, value=h)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

# --- Riga 4: parametri algoritmo per ogni serie ---
ws.cell(row=4, column=1, value="ALGORITMO →")
ws.cell(row=4, column=1).font = Font(bold=True, italic=True, color="666666")
algos = ["AUTO", "SARIMA", "CROSTON", "ENSEMBLE"]
for c, a in enumerate(algos, 3):
    cell = ws.cell(row=4, column=c, value=a)
    cell.font = Font(bold=True, color="C00000")
    cell.alignment = Alignment(horizontal="center")

# --- Dati: 4 serie storiche di 36 mesi ---
np.random.seed(42)
months = [f"{y}-{m:02d}" for y in range(2022, 2025) for m in range(1, 13)]

# S1: Trend crescente
s1 = [round(100 + i*4.5 + np.random.normal(0, 8)) for i in range(36)]
# S2: Stagionale mensile
s2 = [round(150 + 40*np.sin(2*np.pi*i/12) + np.random.normal(0, 10)) for i in range(36)]
# S3: Intermittente
s3_base = [0,0,15,0,0,0,8,0,0,22,0,0, 0,5,0,0,0,18,0,0,0,12,0,0, 0,0,20,0,8,0,0,0,14,0,0,25]
# S4: Puro rumore (difficile da prevedere)
s4 = [round(max(0, 200 + np.random.normal(0, 50))) for i in range(36)]

for row_idx, month in enumerate(months, 5):
    ws.cell(row=row_idx, column=1, value=row_idx - 4)   # SKU_ID progressivo
    ws.cell(row=row_idx, column=2, value=month)
    fill = series_fills[0] if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for c, val in enumerate([s1[row_idx-5], s2[row_idx-5],
                              s3_base[row_idx-5], s4[row_idx-5]], 3):
        cell = ws.cell(row=row_idx, column=c, value=max(0, val))
        cell.border = border

# --- Colonna larghezze ---
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 12
for col in ["C","D","E","F"]:
    ws.column_dimensions[col].width = 16

# --- Foglio Output (vuoto, verrà riempito dal launcher) ---
ws_out = wb.create_sheet("Output")
out_headers = [
    "Serie", "Algoritmo", "Metodo",
    "Forecast_1","Forecast_2","Forecast_3","Forecast_4","Forecast_5",
    "Forecast_6","Forecast_7","Forecast_8","Forecast_9","Forecast_10",
    "Forecast_11","Forecast_12",
    "MASE","lnQ","MAE","RMSE","R2",
    "CI_lower_1","CI_upper_1",
    "Elapsed_sec", "Note"
]
for c, h in enumerate(out_headers, 1):
    cell = ws_out.cell(row=1, column=c, value=h)
    cell.fill   = header_fill
    cell.font   = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = border

ws_out.column_dimensions["A"].width = 10
ws_out.column_dimensions["B"].width = 14
ws_out.column_dimensions["C"].width = 28
for col in [chr(c) for c in range(ord("D"), ord("Z")+1)]:
    ws_out.column_dimensions[col].width = 13

path = "test_data.xlsx"
wb.save(path)
print(f"File creato: {path}")
print("Foglio 'Input'  → contiene 4 serie storiche di 36 mesi")
print("Foglio 'Output' → verrà riempito dal launcher.py")
