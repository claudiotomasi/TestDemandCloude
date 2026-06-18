# =============================================================================
# DGS Forecast Engine — Launcher di Test  v2
# Equivalente a: DGS_forecasting_short.m
#
# USO:
#   python launcher.py                        → usa test_data.xlsx
#   python launcher.py mio_file.xlsx          → usa file specificato
#   python launcher.py mio_file.xlsx SARIMA   → forza algoritmo per tutte
#
# STRUTTURA EXCEL DI INPUT (foglio "Input"):
#   Riga 3 : header colonne
#   Riga 4 : algoritmo per ogni serie (es. AUTO, SARIMA, CROSTON, ...)
#   Righe 5+: dati storici, una colonna per SKU, una riga per periodo
#   Colonna A: indice (ignorata)
#   Colonne B: etichette mese (ignorata)
#   Colonne C+: serie storiche
#
# OUTPUT:
#   Foglio "Output"    — una riga per serie: forecast 12m + KPI
#   Foglio "Confronto" — storia + 4 algoritmi affiancati + grafico per serie
# =============================================================================

import sys, os, time, warnings
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from core.engine import run_forecast

# ── Configurazione ────────────────────────────────────────────────────────────
DEFAULT_EXCEL     = "test_data.xlsx"
DEFAULT_ALGORITHM = "AUTO"
INPUT_SHEET       = "Input"
OUTPUT_SHEET      = "Output"
COMPARE_SHEET     = "Confronto"
HEADER_ROW        = 3
ALGO_ROW          = 4
DATA_START_ROW    = 5
DATA_START_COL    = 3
N_FORECAST        = 12
SEASON_LENGTH     = 12
MASE_PERIOD       = 12
ALGOS_COMPARE     = ["AUTO", "SARIMA", "HW", "ENSEMBLE"]  # 4 algoritmi nel confronto
N_STORIA_CHART    = 12   # ultimi N mesi di storia nel grafico

# ── Stili ─────────────────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", fgColor="2F5496")
FILL_STORIA = PatternFill("solid", fgColor="DEEAF1")
FILL_FC     = [
    PatternFill("solid", fgColor="E2EFDA"),
    PatternFill("solid", fgColor="FFF2CC"),
    PatternFill("solid", fgColor="FCE4D6"),
    PatternFill("solid", fgColor="EAD1DC"),
]
FC_COLORS   = ["375623", "7F6000", "843C0C", "4A235A"]
FILL_BEST   = PatternFill("solid", fgColor="E2EFDA")
FILL_MED    = PatternFill("solid", fgColor="FFF2CC")
FILL_BAD    = PatternFill("solid", fgColor="FCE4D6")
FILL_ALT    = PatternFill("solid", fgColor="F5F5F5")
FONT_HEADER = Font(color="FFFFFF", bold=True)
THIN        = Side(style='thin', color='CCCCCC')
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def mase_fill(mase):
    if mase is None or (isinstance(mase, float) and mase != mase): return FILL_ALT
    return FILL_BEST if mase < 0.3 else (FILL_MED if mase < 0.6 else FILL_BAD)

def mase_font(mase):
    if mase is None or (isinstance(mase, float) and mase != mase): return Font()
    return Font(bold=True, color="375623") if mase < 0.3 else \
           (Font() if mase < 0.6 else Font(color="843C0C"))

def clean(v, decimals=4):
    if v is None: return "N/A"
    try:
        f = float(v)
        return "N/A" if (f != f or abs(f) == float('inf')) else round(f, decimals)
    except Exception:
        return str(v)

def hdr(ws, row, col, value, fill=None, font=None, align="center"):
    cell = ws.cell(row=row, column=col, value=value)
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.alignment = Alignment(horizontal=align)
    cell.border = BORDER
    return cell


# ── Lettura Excel ─────────────────────────────────────────────────────────────

def read_input(path, force_algo=None):
    wb  = openpyxl.load_workbook(path, data_only=True)
    if INPUT_SHEET not in wb.sheetnames:
        raise ValueError(f"Foglio '{INPUT_SHEET}' non trovato in {path}")
    ws  = wb[INPUT_SHEET]

    headers = {}
    for col in ws.iter_cols(min_row=HEADER_ROW, max_row=HEADER_ROW,
                             min_col=DATA_START_COL):
        for cell in col:
            if cell.value is not None:
                headers[cell.column] = str(cell.value)

    algos = {}
    for col_idx in headers:
        cell = ws.cell(row=ALGO_ROW, column=col_idx)
        algos[col_idx] = force_algo.upper() if force_algo else (
            str(cell.value).upper().strip() if cell.value else DEFAULT_ALGORITHM)

    series_list = []
    for col_idx in sorted(headers.keys()):
        values = []
        for row in ws.iter_rows(min_row=DATA_START_ROW,
                                 min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is None: break
                try:    values.append(float(cell.value))
                except: break
        if len(values) < 4:
            print(f"  ⚠ '{headers[col_idx]}': serie troppo corta, saltata")
            continue
        series_list.append({
            'name':      headers[col_idx],
            'algorithm': algos[col_idx],
            'history':   np.array(values, dtype=float),
        })
    wb.close()
    return series_list


# ── Foglio Output ─────────────────────────────────────────────────────────────

def write_output(path, results_out):
    wb = openpyxl.load_workbook(path)
    if OUTPUT_SHEET not in wb.sheetnames: wb.create_sheet(OUTPUT_SHEET)
    ws = wb[OUTPUT_SHEET]
    ws.delete_rows(1, ws.max_row)

    cols = (["Serie", "Algoritmo", "Metodo"] +
            [f"Fc_{i+1}" for i in range(N_FORECAST)] +
            ["MASE","lnQ","MAE","RMSE","R2","CI_L_1","CI_U_1","Elapsed_s","Note"])
    for c, h in enumerate(cols, 1):
        hdr(ws, 1, c, h, FILL_HEADER, FONT_HEADER)

    for row_idx, res in enumerate(results_out, 2):
        kpi  = res.get('kpi', {})
        fc   = res.get('forecast', [])
        lo   = res.get('lower', [])
        hi   = res.get('upper', [])
        mase = kpi.get('MASE')
        note = res.get('fallback', res.get('auto_best', ''))
        fill = mase_fill(mase)
        font = mase_font(mase)
        row_data = (
            [res['name'], res['algorithm'], res.get('method','?')] +
            [clean(v,1) for v in fc] +
            [clean(kpi.get('MASE')), clean(kpi.get('lnQ')),
             clean(kpi.get('MAE'),2), clean(kpi.get('RMSE'),2),
             clean(kpi.get('Rsq')),
             clean(lo[0] if len(lo) > 0 else None,1),
             clean(hi[0] if len(hi) > 0 else None,1),
             clean(res.get('elapsed_sec'),3), note]
        )
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal="center")
            cell.border = BORDER

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    for i in range(N_FORECAST + 9):
        ws.column_dimensions[get_column_letter(4+i)].width = 10

    wb.save(path)
    print(f"  ✔ Foglio '{OUTPUT_SHEET}' scritto")


# ── Foglio Confronto ──────────────────────────────────────────────────────────

def write_comparison(path, series_list):
    from openpyxl.chart import LineChart, Reference, Series as ChartSeries

    wb = openpyxl.load_workbook(path)
    if COMPARE_SHEET in wb.sheetnames: del wb[COMPARE_SHEET]
    ws = wb.create_sheet(COMPARE_SHEET)

    LINE_COLORS  = ["1F4E79", "375623", "7F6000", "843C0C", "4A235A"]
    DASH_STYLES  = ["solid", "dash", "dashDot", "sysDash", "lgDash"]

    current_row = 1
    n_algos     = len(ALGOS_COMPARE)
    total_cols  = 2 + n_algos   # col1=Periodo, col2=Storia, col3..N=Forecast

    # Intestazione colonne fissa
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for i in range(n_algos):
        ws.column_dimensions[get_column_letter(3+i)].width = 14

    for s in series_list:
        name    = s['name']
        history = s['history']
        n       = len(history)

        # --- Titolo serie ---
        title_cell = ws.cell(row=current_row, column=1, value=f"▶  {name}")
        title_cell.font = Font(bold=True, size=12, color="2F5496")
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=total_cols + 2)
        current_row += 1

        # --- Header colonne ---
        hdr(ws, current_row, 1, "Periodo",  FILL_HEADER, FONT_HEADER)
        hdr(ws, current_row, 2, "Storia",   FILL_STORIA,
            Font(bold=True, color="1F4E79"))
        for a_idx, algo in enumerate(ALGOS_COMPARE):
            hdr(ws, current_row, 3+a_idx, f"Fc {algo}",
                FILL_FC[a_idx], Font(bold=True, color=FC_COLORS[a_idx]))
        current_row += 1

        # --- Storia: ultimi N_STORIA_CHART mesi ---
        storia_vals   = history[max(0, n - N_STORIA_CHART):]
        n_st          = len(storia_vals)
        storia_labels = [f"M-{n_st - i}" for i in range(n_st)]
        storia_row_start = current_row

        for i, (label, val) in enumerate(zip(storia_labels, storia_vals)):
            r = current_row + i
            ws.cell(row=r, column=1, value=label).border = BORDER
            cell = ws.cell(row=r, column=2, value=round(float(val), 1))
            cell.fill = FILL_STORIA; cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
            for a_idx in range(n_algos):
                ws.cell(row=r, column=3+a_idx, value=None).border = BORDER

        current_row += n_st

        # --- Forecast: N_FORECAST mesi ---
        fc_row_start = current_row
        fc_data      = {}

        print(f"     Confronto: elaboro {len(ALGOS_COMPARE)} algoritmi per '{name}'...")
        for a_idx, algo in enumerate(ALGOS_COMPARE):
            try:
                r = run_forecast(history, algorithm=algo,
                                 n_forecast=N_FORECAST,
                                 season_length=SEASON_LENGTH,
                                 mase_period=MASE_PERIOD)
                fc_data[algo] = {'forecast': r['forecast'],
                                 'mase': r['kpi'].get('MASE'),
                                 'method': r.get('method', algo)}
            except Exception as e:
                fc_data[algo] = {'forecast': [0]*N_FORECAST,
                                 'mase': None, 'method': f'ERR'}

        for i in range(N_FORECAST):
            r = current_row + i
            ws.cell(row=r, column=1, value=f"Fc+{i+1}").border = BORDER
            ws.cell(row=r, column=2, value=None).border = BORDER
            for a_idx, algo in enumerate(ALGOS_COMPARE):
                val  = fc_data[algo]['forecast'][i] \
                       if i < len(fc_data[algo]['forecast']) else 0
                cell = ws.cell(row=r, column=3+a_idx,
                               value=round(float(val), 1))
                cell.fill = FILL_FC[a_idx]
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="center")

        current_row += N_FORECAST

        # --- Riga MASE ---
        ws.cell(row=current_row, column=1, value="MASE").font = \
            Font(bold=True, italic=True, size=9, color="555555")
        ws.cell(row=current_row, column=2, value="").border = BORDER
        for a_idx, algo in enumerate(ALGOS_COMPARE):
            mase = fc_data[algo]['mase']
            cell = ws.cell(row=current_row, column=3+a_idx, value=clean(mase))
            cell.fill = mase_fill(mase)
            cell.font = Font(bold=True, size=9)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # --- Riga Metodo ---
        ws.cell(row=current_row, column=1, value="Metodo").font = \
            Font(italic=True, size=8, color="888888")
        for a_idx, algo in enumerate(ALGOS_COMPARE):
            cell = ws.cell(row=current_row, column=3+a_idx,
                           value=fc_data[algo]['method'][:22])
            cell.font = Font(size=8, italic=True, color="555555")
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center")
        current_row += 2

        # ── Grafico linee ──────────────────────────────────────────────
        chart = LineChart()
        chart.title  = f"{name} — Storia + Forecast (4 algoritmi)"
        chart.style  = 10
        chart.y_axis.title = "Domanda"
        chart.x_axis.title = "Periodo"
        chart.width  = 22
        chart.height = 13
        chart.legend.position = "b"

        # Serie 1: Storia (linea continua blu scuro, più spessa)
        st_ref  = Reference(ws, min_col=2, max_col=2,
                            min_row=storia_row_start,
                            max_row=storia_row_start + n_st - 1)
        st_ser  = ChartSeries(st_ref, title="Storia")
        st_ser.graphicalProperties.line.solidFill = "1F4E79"
        st_ser.graphicalProperties.line.width     = 25000
        st_ser.smooth = False
        chart.series.append(st_ser)

        # Serie 2..N: Forecast (linee tratteggiate colorate)
        for a_idx, algo in enumerate(ALGOS_COMPARE):
            fc_ref = Reference(ws,
                min_col=3+a_idx, max_col=3+a_idx,
                min_row=fc_row_start,
                max_row=fc_row_start + N_FORECAST - 1)
            mase   = fc_data[algo]['mase']
            label  = f"{algo} (MASE={clean(mase,3)})"
            fc_ser = ChartSeries(fc_ref, title=label)
            fc_ser.graphicalProperties.line.solidFill = FC_COLORS[a_idx % 4]
            fc_ser.graphicalProperties.line.width     = 18000
            # Linea tratteggiata per i forecast
            fc_ser.graphicalProperties.line.dashDot   = "dash"
            fc_ser.smooth = False
            chart.series.append(fc_ser)

        # Etichette asse X (Periodo)
        cat_ref = Reference(ws, min_col=1, max_col=1,
                            min_row=storia_row_start,
                            max_row=fc_row_start + N_FORECAST - 1)
        chart.set_categories(cat_ref)

        # Posizione grafico: a destra della tabella dati
        chart_col  = get_column_letter(total_cols + 2)
        chart_cell = f"{chart_col}{storia_row_start - 1}"
        ws.add_chart(chart, chart_cell)

        # Spazio verticale per il grafico (circa 26 righe Excel)
        current_row = max(current_row, storia_row_start - 1 + 26)
        current_row += 2   # margine tra serie

    wb.save(path)
    print(f"  ✔ Foglio '{COMPARE_SHEET}' scritto con grafici")


# ── Stampa console ────────────────────────────────────────────────────────────

def print_result(name, algo, result):
    kpi     = result.get('kpi', {})
    mase    = kpi.get('MASE')
    lnq     = kpi.get('lnQ')
    fc      = result.get('forecast', [])
    method  = result.get('method', '?')
    elapsed = result.get('elapsed_sec', 0)

    ms   = f"{mase:.4f}" if mase == mase and mase is not None else "  NaN "
    ls   = f"{lnq:.4f}"  if lnq  == lnq  and lnq  is not None else "  NaN "
    icon = "❓" if (mase is None or mase != mase) else \
           ("✅" if mase < 0.3 else ("🟡" if mase < 0.6 else "🔴"))

    print(f"\n  {icon} [{name}]  algo={algo}  metodo={method}")
    print(f"     MASE={ms}  lnQ={ls}  MAE={clean(kpi.get('MAE'),1)}"
          f"  RMSE={clean(kpi.get('RMSE'),1)}  R2={clean(kpi.get('Rsq'),3)}")
    fc_str = "  ".join(str(int(v)) if v==v else "?" for v in fc[:6])
    print(f"     Forecast(1-6): {fc_str} ...  [{elapsed}s]")

    if 'auto_scores' in result:
        best = result.get('auto_best', '?')
        print(f"     AUTO scores:")
        for n, s in result['auto_scores'].items():
            ms2  = f"{s['MASE']:.4f}" if s['MASE']==s['MASE'] else "   NaN"
            flag = " ←" if n == best else ""
            print(f"       {n:<12} MASE={ms2}{flag}")

    if 'weights_used' in result:
        print(f"     Ensemble pesi: " +
              "  ".join(f"{a}={w:.2f}" for a, w in result['weights_used'].items()))

    if 'intermittency_class' in result:
        print(f"     Intermittenza: classe={result['intermittency_class']}"
              f"  ADI={result.get('adi','?')}  CV2={result.get('cv2','?')}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    force_algo = sys.argv[2] if len(sys.argv) > 2 else None

    print("=" * 60)
    print("  DGS Forecast Engine — Launcher di Test  v2")
    print("=" * 60)
    print(f"  File : {excel_path}")
    print(f"  Algo : {force_algo or 'da file (riga 4)'}")
    print()

    if not os.path.exists(excel_path):
        print(f"  ❌ File non trovato: {excel_path}")
        print("  Esegui prima: python create_test_excel.py")
        sys.exit(1)

    print("  Lettura serie da Excel...")
    series_list = read_input(excel_path, force_algo)
    print(f"  Trovate {len(series_list)} serie\n")
    if not series_list:
        print("  Nessuna serie valida."); sys.exit(1)

    # ── Elaborazione principale ──────────────────────────────────────────
    results_out = []
    t_total = time.time()

    for idx, s in enumerate(series_list, 1):
        print(f"  [{idx}/{len(series_list)}] '{s['name']}'"
              f"  ({len(s['history'])} periodi, algo={s['algorithm']})...")
        try:
            result = run_forecast(
                history=s['history'], algorithm=s['algorithm'],
                n_forecast=N_FORECAST, season_length=SEASON_LENGTH,
                mase_period=MASE_PERIOD,
            )
            result['name']      = s['name']
            result['algorithm'] = s['algorithm']
            print_result(s['name'], s['algorithm'], result)
            results_out.append(result)
        except Exception as e:
            print(f"     ❌ Errore: {e}")
            results_out.append({'name': s['name'], 'algorithm': s['algorithm'],
                                'method': 'ERROR', 'forecast': [],
                                'lower': [], 'upper': [], 'kpi': {},
                                'elapsed_sec': 0, 'fallback': str(e)})

    elapsed_total = round(time.time() - t_total, 1)
    print(f"\n{'='*60}")
    print(f"  Elaborazione: {len(series_list)} serie in {elapsed_total}s")

    # ── Scrittura fogli ──────────────────────────────────────────────────
    write_output(excel_path, results_out)

    print()
    print("  Generazione foglio Confronto...")
    write_comparison(excel_path, series_list)

    # ── Riepilogo ────────────────────────────────────────────────────────
    mase_vals = [r['kpi']['MASE'] for r in results_out
                 if r.get('kpi') and r['kpi'].get('MASE') == r['kpi'].get('MASE')]
    if mase_vals:
        print(f"\n  MASE medio : {np.mean(mase_vals):.4f}")
        print(f"  MASE min   : {np.min(mase_vals):.4f}")
        print(f"  MASE max   : {np.max(mase_vals):.4f}")

    print(f"\n  📊 Apri {excel_path} → foglio 'Confronto' per i grafici")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
